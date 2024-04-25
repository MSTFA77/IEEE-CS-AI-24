import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import Workbook, load_workbook
import random
import os

class Student:
    def __init__(self, name, age, grade):
        self.id = random.randint(1000, 9999)  
        self.name = name
        self.age = age
        self.grade = grade

class StudentDB:
    def __init__(self):
        self.students = []

    def add_student(self, name, age, grade):
        self.students.append(Student(name, age, grade))
        self.save_to_excel("students.xlsx")

    def view_students(self):
        return self.students

    def search_student(self, query):
        for student in self.students:
            if query == student.name or str(query) == str(student.id):
                return student
        return None

    def update_student(self, student_id, new_name, new_age, new_grade):
        for student in self.students:
            if student_id == student.id:
                student.name = new_name
                student.age = new_age
                student.grade = new_grade
                self.save_to_excel("students.xlsx")
                return True
        return False

    def delete_student(self, student_id):
        for student in self.students:
            if student_id == student.id:
                self.students.remove(student)
                self.save_to_excel("students.xlsx")
                return True
        return False

    def save_to_excel(self, filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Age", "Grade"])  # Headers
        for student in self.students:
            ws.append([student.id, student.name, student.age, student.grade])
        wb.save(filename)

    def load_from_excel(self, filename):
        if not os.path.exists(filename):
            return False
        self.students = []
        wb = load_workbook(filename)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.students.append(Student(row[1], row[2], row[3]))
        return True

class StudentDBApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Student Database")

        self.db = StudentDB()

        # Define button_width here to make it accessible throughout the class
        self.button_width = 20

        self.create_widgets()

    def create_widgets(self):
        # Customizing colors and fonts
        self.master.config(bg="#f0f0f0")
        self.label = tk.Label(self.master, text="Options:", bg="#f0f0f0", font=("Helvetica", 14))
        self.label.pack()

        # Formatting buttons
        button_bg = "#4caf50"  # Green color for buttons
        button_fg = "white"  # White color for button text
        button_padx = 10
        button_pady = 5

        self.add_button = tk.Button(self.master, text="Add student", command=self.add_student, bg=button_bg, fg=button_fg, width=self.button_width)
        self.add_button.pack(pady=(10, 0))

        self.view_button = tk.Button(self.master, text="View students", command=self.view_students, bg=button_bg, fg=button_fg, width=self.button_width)
        self.view_button.pack(pady=button_pady)

        self.search_button = tk.Button(self.master, text="Search student", command=self.search_student, bg=button_bg, fg=button_fg, width=self.button_width)
        self.search_button.pack(pady=button_pady)

        self.update_button = tk.Button(self.master, text="Update student details", command=self.update_student, bg=button_bg, fg=button_fg, width=self.button_width)
        self.update_button.pack(pady=button_pady)

        self.delete_button = tk.Button(self.master, text="Delete student", command=self.delete_student, bg=button_bg, fg=button_fg, width=self.button_width)
        self.delete_button.pack(pady=button_pady)

        self.load_button = tk.Button(self.master, text="Load from Excel file", command=self.load_from_excel, bg=button_bg, fg=button_fg, width=self.button_width)
        self.load_button.pack(pady=button_pady)

        self.exit_button = tk.Button(self.master, text="Exit", command=self.master.quit, bg="#ff5722", fg=button_fg, width=self.button_width)
        self.exit_button.pack(pady=button_pady)

    def add_student(self):
        add_window = tk.Toplevel(self.master)
        add_window.title("Add Student")
        add_window.config(bg="#f0f0f0")

        tk.Label(add_window, text="Name:", bg="#f0f0f0", font=("Helvetica", 12)).grid(row=0, column=0)
        tk.Label(add_window, text="Age:", bg="#f0f0f0", font=("Helvetica", 12)).grid(row=1, column=0)
        tk.Label(add_window, text="Grade:", bg="#f0f0f0", font=("Helvetica", 12)).grid(row=2, column=0)

        name_entry = tk.Entry(add_window)
        name_entry.grid(row=0, column=1)
        age_entry = tk.Entry(add_window)
        age_entry.grid(row=1, column=1)
        grade_entry = tk.Entry(add_window)
        grade_entry.grid(row=2, column=1)

        def add():
            name = name_entry.get()
            age = int(age_entry.get())
            grade = grade_entry.get()
            self.db.add_student(name, age, grade)
            messagebox.showinfo("Success", "Student added successfully.")
            add_window.destroy()

        tk.Button(add_window, text="Add", command=add, bg="#4caf50", fg="white", width=self.button_width).grid(row=3, columnspan=2, pady=(10, 0))

    def view_students(self):
        view_window = tk.Toplevel(self.master)
        view_window.title("View Students")
        view_window.config(bg="#f0f0f0")

        students = self.db.view_students()

        for i, student in enumerate(students):
            tk.Label(view_window, text=f"ID: {student.id}, Name: {student.name}, Age: {student.age}, Grade: {student.grade}", bg="#f0f0f0", font=("Helvetica", 12)).grid(row=i, column=0)

    def search_student(self):
        search_window = tk.Toplevel(self.master)
        search_window.title("Search Student")
        search_window.config(bg="#f0f0f0")

        tk.Label(search_window, text="Enter student name or ID:", bg="#f0f0f0", font=("Helvetica", 12)).grid(row=0, column=0)

        query_entry = tk.Entry(search_window)
        query_entry.grid(row=0, column=1)

        def search():
            query = query_entry.get()
            student = self.db.search_student(query)
            if student:
                messagebox.showinfo("Result", f"ID: {student.id}, Name: {student.name}, Age: {student.age}, Grade: {student.grade}")
            else:
                messagebox.showinfo("Result", "Student not found.")

        tk.Button(search_window, text="Search", command=search, bg="#4caf50", fg="white", width=self.button_width).grid(row=1, columnspan=2, pady=(10, 0))

    def update_student(self):
        update_window = tk.Toplevel(self.master)
        update_window.title("Update Student Details")
        update_window.config(bg="#f0f0f0")

        tk.Label(update_window, text="Enter student ID:", bg="#f0f0f0", font=("Helvetica", 12)).grid(row=0, column=0)
        tk.Label(update_window, text="New Name:", bg="#f0f0f0", font=("Helvetica", 12)).grid(row=1, column=0)
        tk.Label(update_window, text="New Age:", bg="#f0f0f0", font=("Helvetica", 12)).grid(row=2, column=0)
        tk.Label(update_window, text="New Grade:", bg="#f0f0f0", font=("Helvetica", 12)).grid(row=3, column=0)

        id_entry = tk.Entry(update_window)
        id_entry.grid(row=0, column=1)
        name_entry = tk.Entry(update_window)
        name_entry.grid(row=1, column=1)
        age_entry = tk.Entry(update_window)
        age_entry.grid(row=2, column=1)
        grade_entry = tk.Entry(update_window)
        grade_entry.grid(row=3, column=1)

        def update():
            student_id = int(id_entry.get())
            new_name = name_entry.get()
            new_age = int(age_entry.get())
            new_grade = grade_entry.get()
            if self.db.update_student(student_id, new_name, new_age, new_grade):
                messagebox.showinfo("Success", "Student details updated successfully.")
                update_window.destroy()
            else:
                messagebox.showinfo("Error", "Student not found.")

        tk.Button(update_window, text="Update", command=update, bg="#4caf50", fg="white", width=self.button_width).grid(row=4, columnspan=2, pady=(10, 0))

    def delete_student(self):
        delete_window = tk.Toplevel(self.master)
        delete_window.title("Delete Student")
        delete_window.config(bg="#f0f0f0")

        tk.Label(delete_window, text="Enter student ID:", bg="#f0f0f0", font=("Helvetica", 12)).grid(row=0, column=0)

        id_entry = tk.Entry(delete_window)
        id_entry.grid(row=0, column=1)

        def delete():
            student_id = int(id_entry.get())
            if self.db.delete_student(student_id):
                messagebox.showinfo("Success", "Student deleted successfully.")
                delete_window.destroy()
            else:
                messagebox.showinfo("Error", "Student not found.")

        tk.Button(delete_window, text="Delete", command=delete, bg="#4caf50", fg="white", width=self.button_width).grid(row=1, columnspan=2, pady=(10, 0))

    def load_from_excel(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            if self.db.load_from_excel(filename):
                messagebox.showinfo("Success", "Data loaded from Excel file successfully.")
            else:
                messagebox.showinfo("Error", "File not found or invalid format.")

def main():
    root = tk.Tk()
    app = StudentDBApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
